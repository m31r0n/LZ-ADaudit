"""XLSX workbook builder.

Produces the multi-sheet workbook shipped alongside the HTML report
(Summary, Findings, inventories, evidence, remediation plan, modules,
preflight). IR-specific sheets are not yet emitted — IR data appears
only in the HTML report."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..data import AuditData
from ..utils import (
    fmt_dur as _fmt_dur,
    dfl_label as _dfl_label,
    msdate as _msdate,
    GPO_STATUS,
)
from ..remediation import get_rem as _get_rem, remediation_context as _remediation_context
from ..i18n import t as _t
from ..utils import TIMELINE as _TIMELINE
import re

# Compatibility shim: original body uses module-level names like _XLSX_OK
_XLSX_OK = True


def build_xlsx(data: AuditData, path: Path) -> None:
    if not _XLSX_OK:
        print("  [!] openpyxl not installed — skipping XLSX (pip install openpyxl)")
        return

    wb = openpyxl.Workbook()
    rem_ctx = _remediation_context(data)
    wb.remove(wb.active)  # remove default sheet

    # --- Styles ---
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    DATA_FONT = Font(name="Calibri", size=10)
    MONO_FONT = Font(name="Courier New", size=9)
    FILLS = {
        "high":          PatternFill("solid", fgColor="C00000"),
        "medium":        PatternFill("solid", fgColor="C55A11"),
        "low":           PatternFill("solid", fgColor="2E75B6"),
        "informational": PatternFill("solid", fgColor="808080"),
        "ok":            PatternFill("solid", fgColor="375623"),
        "warn":          PatternFill("solid", fgColor="7F6000"),
        "crit":          PatternFill("solid", fgColor="833030"),
    }
    WHITE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    # Light zebra stripe — avoids the dark-background/black-text readability issue
    ALT_FILL = PatternFill("solid", fgColor="EEF2FF")

    # v1.4.0: Excel limits sheet names to 31 chars and they must be unique. Truncating with [:31]
    # without checking for collision crashes openpyxl. _ws() now disambiguates.
    _used_sheet_names: set[str] = set()

    def _ws(name: str) -> "openpyxl.worksheet.worksheet.Worksheet":
        base = (name or "Sheet")[:31]
        candidate = base
        idx = 2
        while candidate in _used_sheet_names:
            suffix = f"_{idx}"
            candidate = (base[: 31 - len(suffix)] + suffix)
            idx += 1
        _used_sheet_names.add(candidate)
        ws = wb.create_sheet(candidate)
        ws.sheet_view.showGridLines = False
        return ws

    def _hdr(ws: Any, cols: list[str]) -> None:
        for i, c in enumerate(cols, 1):
            cell = ws.cell(1, i, c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    def _row(ws: Any, r: int, values: list[Any], mono: bool = False) -> None:
        fnt = MONO_FONT if mono else DATA_FONT
        fill = ALT_FILL if r % 2 == 0 else None
        for i, v in enumerate(values, 1):
            cell = ws.cell(r, i, v)
            cell.font = fnt
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    def _auto_width(ws: Any, max_w: int = 55) -> None:
        for col in ws.columns:
            w = max(
                len(str(cell.value or "")) for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, max_w)

    def _filter(ws: Any) -> None:
        ws.auto_filter.ref = ws.dimensions

    # ---- Sheet: Summary ----
    ws = _ws("Summary")
    ex = data.execution
    sm = data.summary
    info = [
        ("Tool", ex.get("tool_name", "LZ-ADaudit")),
        ("Version", ex.get("tool_version", "")),
        ("Hostname", ex.get("hostname", "")),
        ("FQDN", ex.get("fqdn", "")),
        ("Domain", ex.get("domain_name", "")),
        ("Forest", ex.get("forest_name", "")),
        ("Profile", ex.get("profile_selected", "")),
        ("Started", ex.get("started_at_utc", "")),
        ("Ended", ex.get("ended_at_utc", "")),
        ("Duration", _fmt_dur(ex.get("duration_seconds"))),
        ("Exit Code", ex.get("exit_code", "")),
    ]
    ws.cell(1, 1, "LZ-ADaudit Audit Summary").font = Font(bold=True, size=14, color="89B4FA")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22
    for i, (k, v) in enumerate(info, 2):
        ws.cell(i, 1, k).font = Font(bold=True, name="Calibri", size=10, color="7F849C")
        ws.cell(i, 2, str(v)).font = DATA_FONT
    r = len(info) + 3
    ws.cell(r, 1, "Findings by Severity").font = Font(bold=True, size=11, color="CDD6F4")
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    for sev, cnt in sm.get("findings_by_severity", {}).items():
        ws.cell(r, 1, sev.upper()).font = WHITE_FONT
        ws.cell(r, 1).fill = FILLS.get(sev, HDR_FILL)
        ws.cell(r, 2, cnt).font = DATA_FONT
        r += 1
    r += 1
    ws.cell(r, 1, "Findings by Category").font = Font(bold=True, size=11, color="CDD6F4")
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    for cat, cnt in sm.get("findings_by_category", {}).items():
        ws.cell(r, 1, cat).font = DATA_FONT
        ws.cell(r, 2, cnt).font = DATA_FONT
        r += 1
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    # ---- Sheet: Findings ----
    ws = _ws("Findings")
    cols = ["Severity", "Score", "Check ID", "Category", "Title",
            "Confidence", "Impact", "Status", "Evidence", "Recommendation", "Detected"]
    _hdr(ws, cols)
    sev_ord = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    for ri, f in enumerate(
        sorted(data.findings, key=lambda x: (sev_ord.get(x.get("severity", ""), 4), -x.get("priority_score", 0))),
        2
    ):
        sev = f.get("severity", "")
        vals = [
            sev.upper(), f.get("priority_score", ""), f.get("check_id", ""),
            f.get("category", ""), f.get("title", ""),
            f.get("confidence", ""), f.get("impact", ""), f.get("status", ""),
            (f.get("evidence", "") or "")[:500],
            f.get("recommendation", ""),
            f.get("created_at_utc", ""),
        ]
        _row(ws, ri, vals)
        ws.cell(ri, 1).fill = FILLS.get(sev, HDR_FILL)
        ws.cell(ri, 1).font = WHITE_FONT
    _filter(ws)
    _auto_width(ws)
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["J"].width = 50

    # ---- Sheet: Domain Info ----
    dom = data.evidence.get("domain", {})
    if dom:
        ws = _ws("Domain Info")
        _hdr(ws, ["Property", "Value"])
        props = [
            ("Domain Name", dom.get("domain_name", "")),
            ("Forest Name", dom.get("forest_name") or ex.get("forest_name", "")),
            ("Domain Functional Level", _dfl_label(dom.get("domain_mode", ""))),
            ("Forest Functional Level", _dfl_label(dom.get("forest_mode", ""))),
            ("PDC Emulator", dom.get("pdc_emulator", "")),
            ("RID Master", dom.get("rid_master", "")),
            ("Infrastructure Master", dom.get("infrastructure_master", "")),
            ("Schema Master", dom.get("schema_master", "")),
            ("Domain Naming Master", dom.get("domain_naming_master", "")),
            ("Recycle Bin Enabled", str(dom.get("recycle_bin_enabled", ""))),
        ]
        for ri, (k, v) in enumerate(props, 2):
            _row(ws, ri, [k, v])
        _auto_width(ws)

    # ---- Sheet: Password Policy ----
    pp = data.evidence.get("password_policy", {}).get("default_policy", {})
    if pp:
        ws = _ws("Password Policy")
        _hdr(ws, ["Setting", "Value"])
        def _td(d: Any) -> str:
            if isinstance(d, dict):
                days = d.get("TotalDays", 0)
                mins = d.get("TotalMinutes", 0)
                if days: return f"{int(days)} days"
                if mins: return f"{int(mins)} min"
            return str(d)
        rows_pp = [
            ("Min Password Length",          pp.get("MinPasswordLength", "")),
            ("Max Password Age",             _td(pp.get("MaxPasswordAge", {}))),
            ("Min Password Age",             _td(pp.get("MinPasswordAge", {}))),
            ("Password History Count",       pp.get("PasswordHistoryCount", "")),
            ("Complexity Enabled",           pp.get("ComplexityEnabled", "")),
            ("Reversible Encryption",        pp.get("ReversibleEncryptionEnabled", "")),
            ("Lockout Threshold",            pp.get("LockoutThreshold", "")),
            ("Lockout Duration",             _td(pp.get("LockoutDuration", {}))),
            ("Lockout Observation Window",   _td(pp.get("LockoutObservationWindow", {}))),
        ]
        for ri, (k, v) in enumerate(rows_pp, 2):
            _row(ws, ri, [k, str(v)])
        _auto_width(ws)

    # ---- Sheet: LDAP Security ----
    ldap = data.evidence.get("ldap", {})
    if ldap or data.txt_files.get("LDAPSecurity"):
        ws = _ws("LDAP Security")
        _hdr(ws, ["Setting", "Raw Value", "Meaning"])
        _INT_M = {0: "No signing — CRITICAL", 1: "Negotiate — not enforced", 2: "Require Signing — OK"}
        _ANON_M = {0: "Anonymous enum allowed — CRITICAL", 1: "No SAM list", 2: "Restricted — OK"}
        _SAM_M  = {0: "Allowed — CRITICAL", 1: "Restricted — OK"}
        ldap_rows = [
            ("LDAPServerIntegrity", ldap.get("LDAPServerIntegrity", ""), _INT_M.get(ldap.get("LDAPServerIntegrity", -1), "?")),
            ("RestrictAnonymous",   ldap.get("RestrictAnonymous", ""),   _ANON_M.get(ldap.get("RestrictAnonymous", -1), "?")),
            ("RestrictAnonymousSam",ldap.get("RestrictAnonymousSam",""), _SAM_M.get(ldap.get("RestrictAnonymousSam", -1), "?")),
        ]
        ri = 2
        for row_vals in ldap_rows:
            _row(ws, ri, list(row_vals)); ri += 1
        for ln in data.txt_files.get("LDAPSecurity", []):
            ws.cell(ri, 1, "Assessment").font = DATA_FONT
            ws.cell(ri, 2, ln).font = MONO_FONT
            ri += 1
        for ln in data.txt_files.get("ntlm_restrictions", []):
            ws.cell(ri, 1, "NTLM GPO").font = DATA_FONT
            ws.cell(ri, 2, ln).font = MONO_FONT
            ri += 1
        for ln in data.txt_files.get("dcs_weak_kerberos_ciphersuite", []):
            parts = ln.split()
            ws.cell(ri, 1, "Weak Kerberos DC").font = DATA_FONT
            ws.cell(ri, 2, parts[0] if parts else ln).font = MONO_FONT
            ws.cell(ri, 3, f"EncType value {parts[1]} (includes RC4)" if len(parts) > 1 else "").font = DATA_FONT
            ri += 1
        _auto_width(ws)

    # ---- Sheet: LAPS ----
    laps = data.evidence.get("laps", {})
    if laps or data.txt_files.get("winlaps_missing-computers"):
        ws = _ws("LAPS")
        _hdr(ws, ["Category", "Value"])
        ri = 2
        if laps:
            for k, v in [("Legacy LAPS Schema", laps.get("legacy_laps_schema")),
                          ("Windows LAPS Schema", laps.get("windows_laps_schema")),
                          ("Domain Functional Level", _dfl_label(laps.get("domain_functional_level", "")))]:
                _row(ws, ri, [k, str(v)]); ri += 1
        for ln in data.txt_files.get("winlaps_missing-computers", []):
            _row(ws, ri, ["Missing LAPS", ln]); ri += 1
        for ln in data.txt_files.get("winlaps_dcs_missing-dsrm", []):
            _row(ws, ri, ["Missing DSRM Password", ln]); ri += 1
        for ln in data.txt_files.get("winlaps_read-extendedrights", []):
            _row(ws, ri, ["Read Extended Rights", ln]); ri += 1
        _auto_width(ws)

    # ---- Sheet: Security Events ----
    if data.security_events_csv:
        ws = _ws("Security Events")
        src = data.security_events_csv
        cols_ev = list(src[0].keys()) if src else []
        _hdr(ws, cols_ev)
        for ri, row_d in enumerate(src, 2):
            _row(ws, ri, [row_d.get(c, "") for c in cols_ev])
        _filter(ws)
        _auto_width(ws)

    # ---- Sheet: Privileged Accounts ----
    def _priv_rows() -> list[tuple[str, str, str]]:
        out: list[tuple[str, str, str]] = []
        def _parse_group(key: str, group: str) -> None:
            for ln in data.txt_files.get(key, []):
                parts = ln.split(" ", 2)
                if len(parts) >= 2:
                    out.append((group, parts[1], parts[2] if len(parts) > 2 else ""))
        def _parse_list(key: str, group: str) -> None:
            for ln in data.txt_files.get(key, []):
                m = re.match(r"^(\S+)\s+\(([^)]+)\)", ln)
                if m: out.append((group, m.group(1), m.group(2)))
                else: out.append((group, ln.strip(), ""))
        _parse_group("domain_admins",         "Domain Admins")
        _parse_group("enterprise_admins",     "Enterprise Admins")
        _parse_group("schema_admins",         "Schema Admins")
        _parse_list( "accounts_userPrivileged","Privileged")
        _parse_list( "accounts_protectedusers","Protected Users")
        return out

    priv = _priv_rows()
    if priv:
        ws = _ws("Privileged Accounts")
        _hdr(ws, ["Group", "SAM Account", "Display Name"])
        for ri, (grp, sam, dn) in enumerate(priv, 2):
            _row(ws, ri, [grp, sam, dn])
        _filter(ws); _auto_width(ws)

    # ---- Sheet: Account Issues ----
    acct_rows: list[tuple[str, str]] = []
    for key, label in [
        ("accounts_inactive",       "Inactive 180d"),
        ("accounts_passdontexpire", "Pwd Never Expires"),
        ("accounts_disabled",       "Disabled"),
        ("new_users",               "New User"),
        ("new_groups",              "New Group"),
    ]:
        for ln in data.txt_files.get(key, []):
            acct_rows.append((label, ln))
    if acct_rows:
        ws = _ws("Account Issues")
        _hdr(ws, ["Issue Type", "Detail"])
        for ri, (t, d) in enumerate(acct_rows, 2):
            _row(ws, ri, [t, d])
        _filter(ws); _auto_width(ws)

    # ---- Sheet: GPOs ----
    gpos_ev = data.evidence.get("gpo", [])
    if gpos_ev:
        ws = _ws("GPOs")
        _hdr(ws, ["Display Name", "Status", "Created", "Modified", "GUID"])
        for ri, g in enumerate(gpos_ev, 2):
            _row(ws, ri, [
                g.get("DisplayName", ""),
                GPO_STATUS.get(g.get("GpoStatus", -1), "?"),
                _msdate(g.get("CreationTime")),
                _msdate(g.get("ModificationTime")),
                g.get("Id", ""),
            ])
        _filter(ws); _auto_width(ws)
        # OU inheritance in same sheet
        ou_lines = data.txt_files.get("ous_inheritedGPOs", [])
        if ou_lines:
            ri = len(gpos_ev) + 4
            ws.cell(ri, 1, "OU GPO Inheritance").font = Font(bold=True, color="89B4FA", size=11)
            ri += 1
            ws.cell(ri, 1, "OU").font = HDR_FONT; ws.cell(ri, 1).fill = HDR_FILL
            ws.cell(ri, 2, "Inherited GPOs").font = HDR_FONT; ws.cell(ri, 2).fill = HDR_FILL
            ri += 1
            for ln in ou_lines:
                if " Inherits these GPOs: " in ln:
                    ou, gpo_list = ln.split(" Inherits these GPOs: ", 1)
                    ws.cell(ri, 1, ou).font = MONO_FONT
                    ws.cell(ri, 2, gpo_list).font = DATA_FONT
                    ri += 1

    # ---- Sheet: ADCS Templates ----
    tpls = data.evidence.get("adcs", {}).get("templates", [])
    if tpls:
        ws = _ws("ADCS Templates")
        _hdr(ws, ["Name", "Display Name", "EKUs"])
        for ri, t in enumerate(tpls, 2):
            _row(ws, ri, [t.get("Name", ""), t.get("displayName", ""), t.get("EKUs", "")])
        _filter(ws); _auto_width(ws)

    # ---- Inventory sheets (distinct names to avoid collision with evidence sheets) ----
    _inv_names = {
        "users": "Users", "computers": "Computers", "groups": "Groups",
        "dcs": "DCs", "service_accounts": "Service Accounts",
        "gpos": "GPO Inventory",       # evidence already has a "GPOs" sheet
        "adcs_templates": "ADCS Inventory",  # evidence already has "ADCS Templates"
    }
    for stem in ["users", "computers", "groups", "dcs", "service_accounts", "gpos", "adcs_templates"]:
        rows_inv = data.inventory.get(stem, [])
        if not rows_inv:
            continue
        ws = _ws(_inv_names.get(stem, stem.replace("_", " ").title())[:31])
        cols_inv = list(rows_inv[0].keys())
        _hdr(ws, cols_inv)
        for ri, row_d in enumerate(rows_inv, 2):
            _row(ws, ri, [row_d.get(c, "") for c in cols_inv])
        _filter(ws); _auto_width(ws)

    # ---- Sheet: Remediation Plan ----
    rem_rows_xl: list[tuple] = []
    seen_xl: set[str] = set()
    sev_ord_xl = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    tl_order_xl = {"immediate": 0, "week": 1, "month": 2, "quarter": 3, "none": 4}
    for f_xl in sorted(
        [f for f in data.findings if f.get("status") != "passed"],
        key=lambda x: (
            tl_order_xl.get(_get_rem(x.get("check_id","")).get("timeline","none"), 4),
            sev_ord_xl.get(x.get("severity",""), 4),
            -x.get("priority_score", 0),
        )
    ):
        cid_xl = f_xl.get("check_id", "")
        if cid_xl in seen_xl:
            continue
        seen_xl.add(cid_xl)
        rem_xl = _get_rem(cid_xl, rem_ctx)
        tl_xl  = rem_xl.get("timeline", "none")
        tl_lbl = _t(f"timeline.{tl_xl}", data.incident.language)
        steps_xl = "\n".join(rem_xl.get("steps", [])) if rem_xl else ""
        refs_xl  = "; ".join(rem_xl.get("references", [])) if rem_xl else ""
        rem_rows_xl.append((
            tl_lbl, f_xl.get("severity","").upper(), cid_xl,
            f_xl.get("title",""), steps_xl, refs_xl,
        ))
    if rem_rows_xl:
        ws = _ws("Remediation Plan")
        _hdr(ws, ["Timeline", "Severity", "Check ID", "Title", "Steps", "References"])
        tl_fills = {
            "Immediate Action": PatternFill("solid", fgColor="C00000"),
            "This Week":        PatternFill("solid", fgColor="C55A11"),
            "This Month":       PatternFill("solid", fgColor="2E75B6"),
            "Next Quarter":     PatternFill("solid", fgColor="375623"),
        }
        for ri, row_vals in enumerate(rem_rows_xl, 2):
            _row(ws, ri, list(row_vals))
            tl_f = tl_fills.get(row_vals[0])
            if tl_f:
                ws.cell(ri, 1).fill = tl_f
                ws.cell(ri, 1).font = WHITE_FONT
            sev_f = FILLS.get(row_vals[1].lower())
            if sev_f:
                ws.cell(ri, 2).fill = sev_f
                ws.cell(ri, 2).font = WHITE_FONT
            ws.cell(ri, 5).alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 45
        _filter(ws)

    # ---- Sheet: Modules ----
    ws = _ws("Modules")
    _hdr(ws, ["Status", "Module", "Display Name", "Findings", "Duration (s)", "Error"])
    for ri, (mod, info) in enumerate(data.execution.get("modules_detail", {}).items(), 2):
        _row(ws, ri, [
            info.get("status", ""),
            mod,
            info.get("display_name", ""),
            info.get("findings_added", 0),
            info.get("duration_seconds", ""),
            info.get("error", "") or "",
        ])
    _filter(ws); _auto_width(ws)

    # ---- Sheet: Preflight ----
    pf_checks = data.preflight.get("checks", [])
    if pf_checks:
        ws = _ws("Preflight")
        _hdr(ws, ["Result", "Check", "Detail"])
        for ri, c in enumerate(pf_checks, 2):
            _row(ws, ri, [c.get("status", "").upper(), c.get("check", ""), c.get("detail", "")])
            st = c.get("status", "")
            fill = FILLS.get("ok" if st == "pass" else ("warn" if st == "warn" else "crit"), None)
            if fill:
                ws.cell(ri, 1).fill = fill
                ws.cell(ri, 1).font = WHITE_FONT
        _auto_width(ws)
